import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
import os
os.chdir("C:\\Users\\Sebastian Pasotr\\Documents\\Data_Coding\\palmtrees")
file = "file.xlsx"
wb = openpyxl.load_workbook(file)
sheet = wb.active

# ---------------------------------------------------------------------- SUMMARY
print(sheet.cell(3,2).value) # [column 3, row 2] == Prices
sheet.max_row # 72
sheet.max_column # 19
## grab a specific cell [col. 3, row 2]
x = sheet.cell(3,2)
print(x.value)
## change value of cell
x.value = int(1000)
x.value

# -------------------------------------------------------------------- SAVE FILE
wb.save("test01.xlsx")
# ---------------------------------------------------------- Packing This Badboy
class readWriteExcel:
    wb = None
    excelFilePath = None

    def __init__(self, excelFilePath):
        self.excelFilePath = excelFilePath
        self.wb = openpyxl.load_workbook(excelFilePath)
    def readByColIndex(self, sheetName,rowIndex,colIndex):
        sheet = self.wb[sheetName]
        return sheet.cell(rowIndex,colIndex).value


driver = webdriver.Firefox()
driver.get("https://apartmentsdavie.com/floor-plans")
html = driver.page_source
soup = BeautifulSoup(html)
avails = list()
for tag in soup.find_all('a'):
    if tag.text == "Check Availability":
        avails.append(tag)
        print(tag.text)
avails[0]


# elem = driver.find_element_by_xpath("/html/body/section[2]/div[2]/div[1]/div/div[2]")
